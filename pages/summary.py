import streamlit as st
import pandas as pd
from datetime import timedelta
from io import BytesIO
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# ================= FULL CSS (UNCHANGED) =================
st.markdown("""
<style>
/* REMOVE default Streamlit page centering */
.block-container {
    padding-top: 1rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}

/* FIX iframe shrinking issue */
iframe {
    width: 100% !important;
    display: block !important;
    margin: 0 !important;
}

/* TABLE container should be full width */
.st-emotion-cache-1kyxreq, .st-emotion-cache-1r6slbn {
    width: 100% !important;
}

/* REMOVE center alignment inside iframe */
html, body {
    margin: 0;
    padding: 0;
    width: 100% !important;
    overflow-x: hidden !important;
}

/* Table itself must be 100% width */
.excel-table {
    width: 100% !important;
    table-layout: fixed !important;
}
</style>
""", unsafe_allow_html=True)

# ================= CSS (last block) =================
st.markdown("""
<style>
.block-container { padding-top: 1rem !important; max-width: 100% !important; }
iframe { width: 100% !important; }
.excel-table { width: 100% !important; table-layout: fixed !important; }
</style>
""", unsafe_allow_html=True)

# ================= NHOST DATABASE CONNECTION =================
@st.cache_resource
def get_db_engine():
    """Create and cache database engine"""
    try:
        db_url = os.getenv("DB_URL")
        if not db_url:
            st.error("DB_URL not found in environment variables")
            return None
        engine = create_engine(db_url)
        return engine
    except Exception as e:
        st.error(f"Database connection failed: {e}")
        return None

def chunk_list(lst, n):
    """Split list into chunks of size n"""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def count_new_upis_for_date(engine, upi_array, cutoff_date):
    """Count new UPIs using the PostgreSQL function"""
    if not upi_array:
        return 0
    
    total_new = 0
    
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(upi_array, 3000):
                query = text("""
                    SELECT missing_count
                    FROM count_new_upi(
                        :p_upi_array,
                        :p_cutoff_date
                    )
                """)
                
                result = conn.execute(query, {
                    "p_upi_array": chunk,
                    "p_cutoff_date": cutoff_date
                }).fetchone()
                
                if result is not None:
                    total_new += result[0]
        
        return total_new
    except Exception as e:
        st.error(f"Error counting new UPIs: {e}")
        return 0

def count_new_banks_for_date(engine, bank_array, cutoff_date):
    """Count new Bank Accounts using the PostgreSQL function"""
    if not bank_array:
        return 0
    
    total_new = 0
    
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(bank_array, 3000):
                query = text("""
                    SELECT missing_count
                    FROM count_new_bank(
                        :p_bank_array,
                        :p_cutoff_date
                    )
                """)
                
                result = conn.execute(query, {
                    "p_bank_array": chunk,
                    "p_cutoff_date": cutoff_date
                }).fetchone()
                
                if result is not None:
                    total_new += result[0]
        
        return total_new
    except Exception as e:
        st.error(f"Error counting new banks: {e}")
        return 0

# ================= UI =================
st.title("UPI, Bank & Website Summary")

uploaded_file = st.file_uploader("Upload Excel or CSV File", type=["xlsx", "xls", "csv"])

if uploaded_file:
    # Get database engine
    engine = get_db_engine()
    if not engine:
        st.error("Cannot proceed without database connection")
        st.stop()
    
    # ---------- LOAD FILE ----------
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)

    df.columns = df.columns.str.strip()
    st.success(f"File Loaded: {uploaded_file.name}")

    # ---------- VALIDATION ----------
    required_cols = [
        "Id", "Feature_type", "Approvd_status", "Input_user",
        "Inserted_date", "Website_url", "Upi_vpa",
        "Bank_account_number", "Search_for", "Upi_bank_account_wallet"
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

    # ---------- FILTER ----------
    filtered_df = df[
        (df["Feature_type"].astype(str).str.strip() == "BS Money Laundering") &
        (df["Approvd_status"].astype(str).str.strip() == "1") &
        (df["Input_user"].astype(str).str.strip().str.lower() != "automated") &
        (df["Search_for"].astype(str).str.strip().isin(["App", "Web"])) &
        (df["Upi_bank_account_wallet"].astype(str).str.strip().isin(["UPI", "Bank Account"]))
    ].copy()

    if filtered_df.empty:
        st.warning("No records found after applying filters.")
        st.stop()

    st.info(f"{len(filtered_df)} rows matched filters")

    # ---------- CLEAN ----------
    def clean_val(x):
        if pd.isna(x):
            return None
        return str(x).strip().lower().replace(" ", "")

    def clean_bank_val(x):
        if pd.isna(x):
            return None
        return str(x).strip()

    filtered_df["Upi_vpa_clean"] = filtered_df["Upi_vpa"].apply(clean_val)
    filtered_df["Bank_acc_clean"] = filtered_df["Bank_account_number"].apply(clean_bank_val)
    filtered_df["Website_url"] = filtered_df["Website_url"].apply(clean_val)
    filtered_df["Inserted_date"] = pd.to_datetime(filtered_df["Inserted_date"], errors="coerce").dt.date

    upi_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip().str.lower() == "upi") &
        (filtered_df["Input_user"].astype(str).str.strip().str.lower() != "automated") &
        (filtered_df["Approvd_status"].astype(str).str.strip() == "1") &
        (filtered_df["Feature_type"].astype(str).str.strip() == "BS Money Laundering")
    ].copy()

    bank_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip() == "Bank Account") &
        (filtered_df["Input_user"].astype(str).str.strip().str.lower() != "automated") &
        (filtered_df["Approvd_status"].astype(str).str.strip() == "1") &
        (filtered_df["Feature_type"].astype(str).str.strip() == "BS Money Laundering")
    ].copy()

    # ---------- GROUPING ----------
    grouped = filtered_df.groupby("Inserted_date").agg(
        website_total=('Id', 'count'),
        Total_UPI=("Upi_vpa_clean", "count"),
        Unique_UPI=("Upi_vpa_clean", pd.Series.nunique),
        unique_website=('Website_url', pd.Series.nunique)
    ).reset_index()

    bank_grouped = bank_df.groupby("Inserted_date").agg(
        Bank_Total=("Bank_acc_clean", "count"),
        Bank_Unique=("Bank_acc_clean", pd.Series.nunique)
    ).reset_index()

    grouped = grouped.merge(bank_grouped, on="Inserted_date", how="left")
    grouped[["Bank_Total", "Bank_Unique"]] = grouped[["Bank_Total", "Bank_Unique"]].fillna(0).astype(int)

    # ---------- SUMMARY (DATE-WISE DB CHECK USING POSTGRESQL FUNCTION) ----------
    summary_data = []

    target_users = [
        "Emp Sunena Yadav",
        "Emp Shubhankar Shukla",
        "Emp Sheetal Dubey"
    ]
    user_rows = []

    # get unique dates first
    all_dates = (
        df["Inserted_date"]
        .pipe(pd.to_datetime, errors="coerce")
        .dt.date
        .dropna()
        .unique()
    )
    freelancer_summary = []

    with st.spinner("Processing data and checking database..."):
        for _, row in grouped.iterrows():
            date = row["Inserted_date"]
            current_date = pd.to_datetime(date).date()
            cutoff_date = (current_date - timedelta(days=1)).strftime("%Y-%m-%d")

            # --- UPI: collect unique UPIs appearing on this date ---
            date_upis = (
                upi_df.loc[upi_df["Inserted_date"] == date, "Upi_vpa_clean"]
                .dropna()
                .astype(str)
                .str.strip()
                .str.lower()
                .unique()
                .tolist()
            )

            # Count new UPIs using PostgreSQL function
            new_upi_today = count_new_upis_for_date(engine, date_upis, cutoff_date)
            print(f"Date: {date}, Unique UPIs: {len(date_upis)}, New UPIs: {new_upi_today}")

            # --- BANK: collect unique bank accounts appearing on this date ---
            date_banks = (
                bank_df.loc[bank_df["Inserted_date"] == date, "Bank_acc_clean"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )

            # Count new Bank Accounts using PostgreSQL function
            new_bank_today = count_new_banks_for_date(engine, date_banks, cutoff_date)
            print(f"Date: {date}, Unique Banks: {len(date_banks)}, New Banks: {new_bank_today}")  

            # --- build summary row ---
            total_upi = int(row["Total_UPI"]) if not pd.isna(row["Total_UPI"]) else 0
            unique_upi = int(row["Unique_UPI"]) if not pd.isna(row["Unique_UPI"]) else 0
            bank_total = int(row["Bank_Total"]) if not pd.isna(row["Bank_Total"]) else 0
            bank_unique = int(row["Bank_Unique"]) if not pd.isna(row["Bank_Unique"]) else 0

            summary_data.append({
                "Date": date,
                "Total": int(row["website_total"]),

                "UPI_Total": total_upi,
                "UPI_Unique": unique_upi,
                "UPI_%": f"{(unique_upi / total_upi * 100):.0f}%" if total_upi else "0%",
                "UPI_New": new_upi_today,
                "UPI_New_%": f"{(new_upi_today / unique_upi * 100):.0f}%" if unique_upi else "0%",

                "Bank_Total": bank_total,
                "Bank_Unique": bank_unique,
                "Bank_%": f"{(bank_unique / bank_total * 100):.0f}%" if bank_total else "0%",
                "Bank_New": new_bank_today,
                "Bank_New_%": f"{(new_bank_today / bank_unique * 100):.0f}%" if bank_unique else "0%",

                "unique_website": int(row["unique_website"]) if not pd.isna(row["unique_website"]) else 0
            })

        # ---------- USER-WISE UPI SUMMARY FOR THIS DATE ----------
        
        for user in target_users:
            user_mask = (
                (upi_df["Inserted_date"] == date) &
                (upi_df["Input_user"].astype(str).str.strip() == user) &
                (upi_df["Approvd_status"].astype(str).str.strip() == "1") &
                (upi_df["Upi_bank_account_wallet"].astype(str).str.strip().str.upper() == "UPI")
            )

            user_df = upi_df.loc[user_mask].copy()

            total = int(len(user_df))
            unique_count = int(user_df["Upi_vpa_clean"].dropna().astype(str).str.strip().nunique())

            # get unique upis for DB check (as list)
            user_upis_list = user_df["Upi_vpa_clean"].dropna().astype(str).str.strip().unique().tolist()

            # call DB function to get new upis for this user's list (uses your chunking function)
            new_count = count_new_upis_for_date(engine, user_upis_list, cutoff_date) if user_upis_list else 0

            unique_pct = f"{(unique_count / total * 100):.0f}%" if total else "0%"
            new_pct = f"{(new_count / unique_count * 100):.0f}%" if unique_count else "0%"

            user_rows.append({
                "Date": date,
                "Input_user": user,
                "Total": total,
                "Unique_UPI_Count": unique_count,
                "Unique_UPI_%": unique_pct,
                "New_UPI_Count": new_count,
                "New_UPI_%": new_pct
            })

        # ---------- FREELANCER SUMMARY FOR THIS DATE ----------
    
        for date in sorted(all_dates):

            cutoff_date = (
                pd.to_datetime(date) - timedelta(days=1)
            ).strftime("%Y-%m-%d")

            # ================= FREELANCER =================
            freelancer_mask = (
                (df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date == date) &
                (df["Input_user"].astype(str).str.contains("Freelancer", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            )

            freelancer_df = df.loc[freelancer_mask].copy()

            # ================= INT (NOT icuser) =================
            int_mask = (
                (df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date == date) &
                (df["Input_user"].astype(str).str.contains("INT", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            )

            int_df = df.loc[int_mask].copy()

            emp_mask = (
                (df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date == date) &
                (df["Input_user"].astype(str).str.contains("Emp", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            )

            emp_df = df.loc[emp_mask].copy()

            # helper function to avoid duplication
            def process_df(date_df):
                if date_df.empty:
                    return 0, 0, 0, 0, 0, 0

                # CLEAN
                date_df["Upi_vpa_clean"] = (
                    date_df["Upi_vpa"]
                    .astype(str).str.strip().str.lower().str.replace(" ", "")
                )

                date_df["Bank_acc_clean"] = (
                    date_df["Bank_account_number"].apply(clean_bank_val)
                )

                # ---------- UPI ----------
                upi_df = date_df[
                    date_df["Upi_bank_account_wallet"]
                    .astype(str).str.strip().str.upper() == "UPI"
                ]

                total_upi = len(upi_df)
                unique_upi_list = upi_df["Upi_vpa_clean"].dropna().unique().tolist()
                unique_upi = len(unique_upi_list)

                new_upi = (
                    count_new_upis_for_date(engine, unique_upi_list, cutoff_date)
                    if unique_upi_list else 0
                )

                # ---------- BANK ----------
                bank_df = date_df[
                    date_df["Upi_bank_account_wallet"]
                    .astype(str).str.strip() == "Bank Account"
                ]

                total_bank = len(bank_df)
                unique_bank_list = bank_df["Bank_acc_clean"].dropna().unique().tolist()
                unique_bank = len(unique_bank_list)

                new_bank = (
                    count_new_banks_for_date(engine, unique_bank_list, cutoff_date)
                    if unique_bank_list else 0
                )

                return total_upi, unique_upi, new_upi, total_bank, unique_bank, new_bank

            # ---------- PROCESS BOTH ----------
            f_total_upi, f_unique_upi, f_new_upi, f_total_bank, f_unique_bank, f_new_bank = (
                process_df(freelancer_df)
            )

            i_total_upi, i_unique_upi, i_new_upi, i_total_bank, i_unique_bank, i_new_bank = (
                process_df(int_df)
            )

            e_total_upi, e_unique_upi, e_new_upi, e_total_bank, e_unique_bank, e_new_bank = (
                process_df(emp_df)
            )

            # ---------- APPEND ----------
            freelancer_summary.append({
                "User_Type": "Freelancer",
                "Date": date,
                "Total_UPI": f_total_upi,
                "Unique_UPI": f_unique_upi,
                "New_UPI": f_new_upi,
                "Total_Bank": f_total_bank,
                "Unique_Bank": f_unique_bank,
                "New_Bank": f_new_bank
            })

            freelancer_summary.append({
                "User_Type": "INT",
                "Date": date,
                "Total_UPI": i_total_upi,
                "Unique_UPI": i_unique_upi,
                "New_UPI": i_new_upi,
                "Total_Bank": i_total_bank,
                "Unique_Bank": i_unique_bank,
                "New_Bank": i_new_bank
            })

            freelancer_summary.append({
                "User_Type": "Employee",
                "Date": date,
                "Total_UPI": e_total_upi,
                "Unique_UPI": e_unique_upi,
                "New_UPI": e_new_upi,
                "Total_Bank": e_total_bank,
                "Unique_Bank": e_unique_bank,
                "New_Bank": e_new_bank
            })


    summary_df = pd.DataFrame(summary_data)
    multiple_summary_df = pd.DataFrame(user_rows)
    freelancer_summary_df = pd.DataFrame(freelancer_summary)

    # ---------- DISPLAY ----------
    st.subheader("📊 Summary Report")

    summary_type = st.selectbox(
        "Select Summary Type",
        [
            "UPI & Bank Summary",
            "Multiple User's Summary",
            "Intern & Freelancer Summary"
        ]
    )

    if summary_type == "UPI & Bank Summary":         
        # Create styled HTML table
        html_table = """
            <style>
            .table-container {
                width: 100%;
                overflow-x: auto;
                margin: 0;
                padding: 0;
            }

            .excel-table {
                border-collapse: collapse;
                font-family: 'Segoe UI', sans-serif;
                font-size: 13px;
                width: 100% !important;
                table-layout: fixed !important;
                min-width: unset;
            }

            .excel-table th, .excel-table td {
                border: 1px solid #ccc;
                text-align: center;
                padding: 6px 4px !important;
                white-space: normal !important;
                word-wrap: break-word !important;
                overflow: hidden;
            }

            .excel-table th:nth-child(1), .excel-table td:nth-child(1) { width: 8%; }
            .excel-table th:nth-child(2), .excel-table td:nth-child(2) { width: 5%; }
            .excel-table th:nth-child(3), .excel-table td:nth-child(3) { width: 5%; }
            .excel-table th:nth-child(4), .excel-table td:nth-child(4) { width: 6%; }
            .excel-table th:nth-child(5), .excel-table td:nth-child(5) { width: 4%; }
            .excel-table th:nth-child(6), .excel-table td:nth-child(6) { width: 5%; }
            .excel-table th:nth-child(7), .excel-table td:nth-child(7) { width: 4%; }
            .excel-table th:nth-child(8), .excel-table td:nth-child(8) { width: 7%; }
            .excel-table th:nth-child(9), .excel-table td:nth-child(9) { width: 6%; }
            .excel-table th:nth-child(10), .excel-table td:nth-child(10) { width: 4%; }
            .excel-table th:nth-child(11), .excel-table td:nth-child(11) { width: 5%; }
            .excel-table th:nth-child(12), .excel-table td:nth-child(12) { width: 4%; }
            .excel-table th:nth-child(13), .excel-table td:nth-child(13) { width: 8%; }

            .excel-table thead tr:first-child th {
                background-color: #cbd5e1;
                font-size: 16px;
                font-weight: 700;
                padding: 8px;
            }

            .excel-table thead tr:nth-child(2) th {
                background-color: #cbd5e1;
                font-size: 14px;
                font-weight: 600;
            }

            .excel-table thead tr:nth-child(3) th {
                background-color: #e2e8f0;
                font-weight: 500;
                font-size: 12px;
            }

            .excel-table td {
                background-color: #f8fafc;
            }
            </style>

            <div class="table-container">
            <table class="excel-table">
                <thead>
                    <tr>
                        <th colspan="13">UPI, Bank & Website Report</th>
                    </tr>

                    <tr>
                        <th rowspan="2">Date</th>
                        <th rowspan="2">Total</th>
                        <th colspan="5">UPI</th>
                        <th colspan="5">Bank</th>
                        <th rowspan="2">Unique Website</th>
                    </tr>

                    <tr>
                        <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
                        <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
                    </tr>
                </thead>

                <tbody>
        """

        for _, row in summary_df.iterrows():
            html_table += f"""
                <tr>
                    <td>{row['Date']}</td>
                    <td>{row['Total']}</td>
                    <td>{row['UPI_Total']}</td>
                    <td>{row['UPI_Unique']}</td>
                    <td>{row['UPI_%']}</td>
                    <td>{row['UPI_New']}</td>
                    <td>{row['UPI_New_%']}</td>
                    <td>{row['Bank_Total']}</td>
                    <td>{row['Bank_Unique']}</td>
                    <td>{row['Bank_%']}</td>
                    <td>{row['Bank_New']}</td>
                    <td>{row['Bank_New_%']}</td>
                    <td>{row['unique_website']}</td>
                </tr>
            """

        html_table += "</tbody></table></div>"

        # Render the HTML
        components.html(
            html_table,
            height=450,
            scrolling=True
        )

    elif summary_type == "Multiple User's Summary":
        # prepare HTML table
        multiple_user_table = f"""
            <style>
            .table-user {{
                width:100%;
                border-collapse:collapse;
                font-family:'Segoe UI', sans-serif;
                font-size:14px;
            }}
            .table-user th, .table-user td {{
                border:1px solid #000;
                padding:6px 10px;
                text-align:center;
            }}
            .table-user thead th {{
                background:#cfe8b0;
                font-weight:700;
            }}
            .table-user tfoot td {{
                font-weight:700;
                background:#ffffff;
            }}
            .table-user td.name {{
                text-align:left;
            }}
            </style>

            <div style="margin-top:12px;">
            <table class="table-user">
                <thead>
                    <tr>
                        <th colspan="6">Multiple User's Counts ({date})</th>
                    </tr>
                    <tr>
                        <th rowspan="2">Name</th>
                        <th rowspan="2">Total</th>
                        <th colspan="2">Unique UPI</th>
                        <th colspan="2">New UPI</th>
                    </tr>
                    <tr>
                        <th>Count</th><th>%</th>
                        <th>Count</th><th>%</th>
                    </tr>
                </thead>
            <tbody>
            """
        
        for _, row in multiple_summary_df.iterrows():
            multiple_user_table += f"""
                <tr>
                    <td>{row['Input_user']}</td>
                    <td>{row['Total']}</td>
                    <td>{row['Unique_UPI_Count']}</td>
                    <td>{row['Unique_UPI_%']}</td>
                    <td>{row['New_UPI_Count']}</td>
                    <td>{row['New_UPI_%']}</td>
                </tr>
            """

        # Render the HTML
        components.html(
            multiple_user_table,
            height=450,
            scrolling=True
        )
        
    elif summary_type == "Intern & Freelancer Summary":
        # prepare HTML table
        freelancer_table = f"""
            <style>
            .table-user {{
                width:100%;
                border-collapse:collapse;
                font-family:'Segoe UI', sans-serif;
                font-size:14px;
            }}
            .table-user th, .table-user td {{
                border:1px solid #000;
                padding:6px 10px;
                text-align:center;
            }}
            .table-user thead th {{
                background:#cfe8b0;
                font-weight:700;
            }}
            .table-user tfoot td {{
                font-weight:700;
                background:#ffffff;
            }}
            .table-user td.name {{
                text-align:left;
            }}
            </style>

            <div style="margin-top:12px;">
            <table class="table-user">
                <thead>
                    <tr><th colspan="8">Employee, Intern & Freelancer Summary</th></tr>
                    <tr>
                        <th>User</th>
                        <th>Date</th>
                        <th>Total UPI</th>
                        <th>Unique UPI</th>
                        <th>New UPI</th>
                        <th>Total Bank</th>
                        <th>Unique Bank</th>
                        <th>New Bank</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>Employee</td>
                        <td rowspan="3">{date}</td>
                        <td>{e_total_upi}</td>
                        <td>{e_unique_upi}</td>
                        <td>{e_new_upi}</td>
                        <td>{e_total_bank}</td>
                        <td>{e_unique_bank}</td>
                        <td>{e_new_bank}</td>
                    </tr>
                    <tr>
                        <td>Intern</td>
                        <td>{i_total_upi}</td>
                        <td>{i_unique_upi}</td>
                        <td>{i_new_upi}</td>
                        <td>{i_total_bank}</td>
                        <td>{i_unique_bank}</td>
                        <td>{i_new_bank}</td>
                    </tr>
                    <tr>
                        <td>Freelancer</td>
                        <td>{f_total_upi}</td>
                        <td>{f_unique_upi}</td>
                        <td>{f_new_upi}</td>
                        <td>{f_total_bank}</td>
                        <td>{f_unique_bank}</td>
                        <td>{f_new_bank}</td>
                    </tr>
                </tbody>
            </table>
            """

        # Render the HTML
        components.html(
            freelancer_table,
            height=450,
            scrolling=True
        )

    # ---------- EXCEL EXPORT ----------
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(summary_df.columns.tolist())
    for r in dataframe_to_rows(summary_df, index=False, header=False):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(output)
    output.seek(0)

    st.download_button(
        "Download Summary Excel",
        data=output,
        file_name="upi_bank_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("📤 Please upload a file to generate the report.")