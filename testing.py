import streamlit as st
import pandas as pd
from datetime import timedelta
from io import BytesIO
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components
import os
from dotenv import load_dotenv

load_dotenv()

st.markdown("""
<style>
.block-container {
    padding-top: 1rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}
iframe { width: 100% !important; display: block !important; margin: 0 !important; }
html, body { margin: 0; padding: 0; width: 100% !important; overflow-x: hidden !important; }
.excel-table { width: 100% !important; table-layout: fixed !important; }
</style>
""", unsafe_allow_html=True)


# ================= HELPER FUNCTIONS =================
def find_column(cols, keys):
    for c in cols:
        cc = c.lower().replace(" ", "").replace("_", "")
        for k in keys:
            if k in cc:
                return c
    return None


@st.cache_resource
def get_db_engine():
    try:
        db_url = os.getenv("DB_URL")
        if not db_url:
            st.error("DB_URL not found in environment variables")
            return None
        return create_engine(db_url)
    except Exception as e:
        st.error(f"Database connection failed: {e}")
        return None


def chunk_list(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def count_new_upis_for_date(engine, upi_array, cutoff_date):
    if not upi_array:
        return 0
    total_new = 0
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(upi_array, 3000):
                result = conn.execute(
                    text("SELECT missing_count FROM count_new_upi(:p_upi_array, :p_cutoff_date)"),
                    {"p_upi_array": chunk, "p_cutoff_date": cutoff_date}
                ).fetchone()
                if result is not None:
                    total_new += result[0]
        return total_new
    except Exception as e:
        st.error(f"Error counting new UPIs: {e}")
        return 0


def count_new_banks_for_date(engine, bank_array, cutoff_date):
    if not bank_array:
        return 0
    total_new = 0
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(bank_array, 3000):
                result = conn.execute(
                    text("SELECT missing_count FROM count_new_bank(:p_bank_array, :p_cutoff_date)"),
                    {"p_bank_array": chunk, "p_cutoff_date": cutoff_date}
                ).fetchone()
                if result is not None:
                    total_new += result[0]
        return total_new
    except Exception as e:
        st.error(f"Error counting new banks: {e}")
        return 0


def clean_val(x):
    if pd.isna(x):
        return None
    return str(x).strip().lower().replace(" ", "")


def clean_bank_val(x):
    if pd.isna(x):
        return None
    return str(x).strip()


def process_df(date_df, engine, cutoff_date):
    """Helper to compute UPI & Bank stats for a given sub-dataframe."""
    if date_df.empty:
        return 0, 0, 0, 0, 0, 0

    date_df = date_df.copy()
    date_df["Upi_vpa_clean"] = date_df["Upi_vpa"].astype(str).str.strip().str.lower().str.replace(" ", "")
    date_df["Bank_acc_clean"] = date_df["Bank_account_number"].apply(clean_bank_val)

    upi_sub = date_df[date_df["Upi_bank_account_wallet"].astype(str).str.strip().str.upper() == "UPI"]
    total_upi = len(upi_sub)
    unique_upi_list = upi_sub["Upi_vpa_clean"].dropna().unique().tolist()
    unique_upi = len(unique_upi_list)
    new_upi = count_new_upis_for_date(engine, unique_upi_list, cutoff_date) if unique_upi_list else 0

    bank_sub = date_df[date_df["Upi_bank_account_wallet"].astype(str).str.strip() == "Bank Account"]
    total_bank = len(bank_sub)
    unique_bank_list = bank_sub["Bank_acc_clean"].dropna().unique().tolist()
    unique_bank = len(unique_bank_list)
    new_bank = count_new_banks_for_date(engine, unique_bank_list, cutoff_date) if unique_bank_list else 0

    return total_upi, unique_upi, new_upi, total_bank, unique_bank, new_bank


# ================= EXCEL EXPORT =================
def build_excel(summary_df, multiple_summary_df, freelancer_summary_df):
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="CBD5E1")
    subheader_fill = PatternFill("solid", fgColor="E2E8F0")
    green_fill = PatternFill("solid", fgColor="CFE8B0")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    def set_header(ws, row, col, value, rowspan=1, colspan=1, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = bold
        cell.alignment = center
        if fill:
            cell.fill = fill
        if rowspan > 1 or colspan > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row + rowspan - 1, end_column=col + colspan - 1
            )

    # ============ SHEET 1: UPI & Bank Summary ============
    ws1 = wb.active
    ws1.title = "UPI & Bank Summary"

    set_header(ws1, 1, 1, "UPI, Bank & Website Report", colspan=13, fill=header_fill)
    set_header(ws1, 2, 1, "Date", rowspan=2, fill=header_fill)
    set_header(ws1, 2, 2, "Total", rowspan=2, fill=header_fill)
    set_header(ws1, 2, 3, "UPI", colspan=5, fill=header_fill)
    set_header(ws1, 2, 8, "Bank", colspan=5, fill=header_fill)
    set_header(ws1, 2, 13, "Unique Website", rowspan=2, fill=header_fill)

    for col, label in enumerate(["Total", "Unique", "%", "New", "%"], start=3):
        set_header(ws1, 3, col, label, fill=subheader_fill)
    for col, label in enumerate(["Total", "Unique", "%", "New", "%"], start=8):
        set_header(ws1, 3, col, label, fill=subheader_fill)

    for r, row in enumerate(summary_df.itertuples(), start=4):
        ws1.append([
            str(row.Date), row.Total,
            row.UPI_Total, row.UPI_Unique, row.UPI_pct, row.UPI_New, row.UPI_New_pct,
            row.Bank_Total, row.Bank_Unique, row.Bank_pct, row.Bank_New, row.Bank_New_pct,
            row.unique_website
        ])
        for col in range(1, 14):
            ws1.cell(row=r, column=col).alignment = center

    for col_idx, width in enumerate([12, 7, 7, 7, 6, 6, 7, 7, 7, 6, 6, 7, 10], start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ============ SHEET 2: Multiple User Summary ============
    ws2 = wb.create_sheet("Multiple User Summary")

    dates = multiple_summary_df["Date"].unique()
    date_label = str(dates[-1]) if len(dates) > 0 else "N/A"

    set_header(ws2, 1, 1, f"Multiple User's Counts ({date_label})", colspan=6, fill=green_fill)
    set_header(ws2, 2, 1, "Name", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 2, "Total", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 3, "Unique UPI", colspan=2, fill=green_fill)
    set_header(ws2, 2, 5, "New UPI", colspan=2, fill=green_fill)
    for col, label in enumerate(["Count", "%", "Count", "%"], start=3):
        set_header(ws2, 3, col, label, fill=green_fill)

    for r, row in enumerate(multiple_summary_df.itertuples(), start=4):
        ws2.append([
            row.Input_user, row.Total,
            row.Unique_UPI_Count, row.Unique_UPI_pct,
            row.New_UPI_Count, row.New_UPI_pct
        ])
        for col in range(1, 7):
            ws2.cell(row=r, column=col).alignment = center

    # Totals row
    total_row = r + 1
    total_total = multiple_summary_df["Total"].sum()
    total_unique = multiple_summary_df["Unique_UPI_Count"].sum()
    total_new = multiple_summary_df["New_UPI_Count"].sum()
    total_unique_pct = f"{(total_unique / total_total * 100):.0f}%" if total_total else "0%"
    total_new_pct = f"{(total_new / total_unique * 100):.0f}%" if total_unique else "0%"

    ws2.append(["Total", total_total, total_unique, total_unique_pct, total_new, total_new_pct])
    for col in range(1, 7):
        cell = ws2.cell(row=total_row, column=col)
        cell.font = bold
        cell.fill = green_fill
        cell.alignment = center

    for col_idx, width in enumerate([25, 8, 12, 8, 12, 8], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ============ SHEET 3: Employee, Intern & Freelancer ============
    ws3 = wb.create_sheet("Emp Intern Freelancer")

    set_header(ws3, 1, 1, "Employee, Intern & Freelancer Summary", colspan=8, fill=green_fill)
    set_header(ws3, 2, 1, "User", rowspan=2, fill=green_fill)
    set_header(ws3, 2, 2, "Date", rowspan=2, fill=green_fill)
    set_header(ws3, 2, 3, "UPI", colspan=3, fill=green_fill)
    set_header(ws3, 2, 6, "Bank", colspan=3, fill=green_fill)
    for col, label in enumerate(["Total UPI", "Unique UPI", "New UPI"], start=3):
        set_header(ws3, 3, col, label, fill=green_fill)
    for col, label in enumerate(["Total Bank", "Unique Bank", "New Bank"], start=6):
        set_header(ws3, 3, col, label, fill=green_fill)

    row_idx = 4
    for date in sorted(freelancer_summary_df["Date"].unique()):
        day_df = freelancer_summary_df[freelancer_summary_df["Date"] == date]
        for user_type in ["Employee", "INT", "Freelancer"]:
            r = day_df[day_df["User_Type"] == user_type]
            if not r.empty:
                r = r.iloc[0]
                ws3.append([
                    user_type, str(date),
                    r["Total_UPI"], r["Unique_UPI"], r["New_UPI"],
                    r["Total_Bank"], r["Unique_Bank"], r["New_Bank"]
                ])
            else:
                ws3.append([user_type, str(date), 0, 0, 0, 0, 0, 0])
            for col in range(1, 9):
                ws3.cell(row=row_idx, column=col).alignment = center
            row_idx += 1

        # Totals per date
        totals = day_df.sum(numeric_only=True)
        ws3.append([
            "Total", "NA",
            int(totals.get("Total_UPI", 0)), int(totals.get("Unique_UPI", 0)), int(totals.get("New_UPI", 0)),
            int(totals.get("Total_Bank", 0)), int(totals.get("Unique_Bank", 0)), int(totals.get("New_Bank", 0))
        ])
        for col in range(1, 9):
            cell = ws3.cell(row=row_idx, column=col)
            cell.font = bold
            cell.alignment = center
        row_idx += 1

    for col_idx, width in enumerate([12, 12, 10, 10, 10, 10, 12, 10], start=1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ================= UI =================
st.title("UPI, Bank & Website Summary")

uploaded_file = st.file_uploader("Upload Excel or CSV File", type=["xlsx", "xls", "csv"])

if uploaded_file:
    engine = get_db_engine()
    if not engine:
        st.error("Cannot proceed without database connection")
        st.stop()

    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)

    df.columns = df.columns.str.strip()
    st.success(f"File Loaded: {uploaded_file.name}")

    required_cols = [
        "Id", "Feature_type", "Approvd_status", "Input_user",
        "Inserted_date", "Website_url", "Upi_vpa",
        "Bank_account_number", "Search_for", "Upi_bank_account_wallet"
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

    filtered_df = df[
        (df["Feature_type"].astype(str).str.strip() == "BS Money Laundering") &
        (df["Approvd_status"].astype(str).str.strip() == "1") &
        (df["Input_user"].astype(str).str.strip().str.lower() != "automated") &
        (df["Search_for"].astype(str).str.strip().isin(["App", "Web"])) &
        (df["Upi_bank_account_wallet"].astype(str).str.strip().isin(["UPI", "Bank Account"]))
    ].copy()

    if filtered_df.empty:
        st.warning("No records found after applying filters.")
        st.stop()

    st.info(f"{len(filtered_df)} rows matched filters")

    filtered_df["Upi_vpa_clean"] = filtered_df["Upi_vpa"].apply(clean_val)
    filtered_df["Bank_acc_clean"] = filtered_df["Bank_account_number"].apply(clean_bank_val)
    filtered_df["Website_url"] = filtered_df["Website_url"].apply(clean_val)
    filtered_df["Inserted_date"] = pd.to_datetime(filtered_df["Inserted_date"], errors="coerce").dt.date

    upi_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip().str.lower() == "upi")
    ].copy()

    bank_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip() == "Bank Account")
    ].copy()

    grouped = filtered_df.groupby("Inserted_date").agg(
        website_total=('Id', 'count'),
        Total_UPI=("Upi_vpa_clean", "count"),
        Unique_UPI=("Upi_vpa_clean", pd.Series.nunique),
        unique_website=('Website_url', pd.Series.nunique)
    ).reset_index()

    bank_grouped = bank_df.groupby("Inserted_date").agg(
        Bank_Total=("Bank_acc_clean", "count"),
        Bank_Unique=("Bank_acc_clean", pd.Series.nunique)
    ).reset_index()

    grouped = grouped.merge(bank_grouped, on="Inserted_date", how="left")
    grouped[["Bank_Total", "Bank_Unique"]] = grouped[["Bank_Total", "Bank_Unique"]].fillna(0).astype(int)

    all_dates = (
        df["Inserted_date"]
        .pipe(pd.to_datetime, errors="coerce")
        .dt.date
        .dropna()
        .unique()
    )

    target_users = [
        "Emp Sunena Yadav",
        "Emp Shubhankar Shukla",
        "Emp Sheetal Dubey"
    ]

    summary_data = []
    user_rows = []        # FIX 1: collect across ALL dates (was only last date)
    freelancer_summary = []

    with st.spinner("Processing data and checking database..."):
        for _, row in grouped.iterrows():
            date = row["Inserted_date"]
            cutoff_date = (pd.to_datetime(date) - timedelta(days=1)).strftime("%Y-%m-%d")

            # --- UPI new count ---
            date_upis = (
                upi_df.loc[upi_df["Inserted_date"] == date, "Upi_vpa_clean"]
                .dropna().astype(str).str.strip().str.lower().unique().tolist()
            )
            new_upi_today = count_new_upis_for_date(engine, date_upis, cutoff_date)

            # --- Bank new count ---
            date_banks = (
                bank_df.loc[bank_df["Inserted_date"] == date, "Bank_acc_clean"]
                .dropna().astype(str).str.strip().unique().tolist()
            )
            new_bank_today = count_new_banks_for_date(engine, date_banks, cutoff_date)

            total_upi = int(row["Total_UPI"]) if not pd.isna(row["Total_UPI"]) else 0
            unique_upi = int(row["Unique_UPI"]) if not pd.isna(row["Unique_UPI"]) else 0
            bank_total = int(row["Bank_Total"]) if not pd.isna(row["Bank_Total"]) else 0
            bank_unique = int(row["Bank_Unique"]) if not pd.isna(row["Bank_Unique"]) else 0

            summary_data.append({
                "Date": date,
                "Total": int(row["website_total"]),
                "UPI_Total": total_upi,
                "UPI_Unique": unique_upi,
                "UPI_pct": f"{(unique_upi / total_upi * 100):.0f}%" if total_upi else "0%",
                "UPI_New": new_upi_today,
                "UPI_New_pct": f"{(new_upi_today / unique_upi * 100):.0f}%" if unique_upi else "0%",
                "Bank_Total": bank_total,
                "Bank_Unique": bank_unique,
                "Bank_pct": f"{(bank_unique / bank_total * 100):.0f}%" if bank_total else "0%",
                "Bank_New": new_bank_today,
                "Bank_New_pct": f"{(new_bank_today / bank_unique * 100):.0f}%" if bank_unique else "0%",
                "unique_website": int(row["unique_website"]) if not pd.isna(row["unique_website"]) else 0
            })

            # FIX 1: Multiple user summary — now runs for EVERY date
            for user in target_users:
                user_mask = (
                    (upi_df["Inserted_date"] == date) &
                    (upi_df["Input_user"].astype(str).str.strip() == user)
                )
                user_sub = upi_df.loc[user_mask].copy()
                total = int(len(user_sub))
                unique_count = int(user_sub["Upi_vpa_clean"].dropna().astype(str).str.strip().nunique())
                user_upis_list = user_sub["Upi_vpa_clean"].dropna().astype(str).str.strip().unique().tolist()
                new_count = count_new_upis_for_date(engine, user_upis_list, cutoff_date) if user_upis_list else 0

                user_rows.append({
                    "Date": date,
                    "Input_user": user,
                    "Total": total,
                    "Unique_UPI_Count": unique_count,
                    "Unique_UPI_pct": f"{(unique_count / total * 100):.0f}%" if total else "0%",
                    "New_UPI_Count": new_count,
                    "New_UPI_pct": f"{(new_count / unique_count * 100):.0f}%" if unique_count else "0%"
                })

        # FIX 2: Freelancer loop — was correct structure but reuse same cutoff_date
        for date in sorted(all_dates):
            cutoff_date = (pd.to_datetime(date) - timedelta(days=1)).strftime("%Y-%m-%d")

            date_series = df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date

            freelancer_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("Freelancer", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            int_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("INT", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            emp_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("Emp", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            for user_type, sub_df in [("Freelancer", freelancer_df), ("INT", int_df), ("Employee", emp_df)]:
                t_upi, u_upi, n_upi, t_bank, u_bank, n_bank = process_df(sub_df, engine, cutoff_date)
                freelancer_summary.append({
                    "User_Type": user_type,
                    "Date": date,
                    "Total_UPI": t_upi,
                    "Unique_UPI": u_upi,
                    "New_UPI": n_upi,
                    "Total_Bank": t_bank,
                    "Unique_Bank": u_bank,
                    "New_Bank": n_bank
                })

    summary_df = pd.DataFrame(summary_data)
    multiple_summary_df = pd.DataFrame(user_rows)
    freelancer_summary_df = pd.DataFrame(freelancer_summary)

    # ================= DISPLAY =================
    st.subheader("📊 Summary Report")

    summary_type = st.selectbox(
        "Select Summary Type",
        ["UPI & Bank Summary", "Multiple User's Summary", "Employee, Intern & Freelancer Summary"]
    )

    if summary_type == "UPI & Bank Summary":
        html_table = """
        <style>
        .table-container { width:100%; overflow-x:auto; }
        .excel-table { border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:13px; width:100% !important; table-layout:fixed !important; }
        .excel-table th, .excel-table td { border:1px solid #ccc; text-align:center; padding:6px 4px; white-space:normal; word-wrap:break-word; }
        .excel-table thead tr:first-child th { background:#cbd5e1; font-size:16px; font-weight:700; }
        .excel-table thead tr:nth-child(2) th { background:#cbd5e1; font-size:14px; font-weight:600; }
        .excel-table thead tr:nth-child(3) th { background:#e2e8f0; font-size:12px; }
        .excel-table td { background:#f8fafc; }
        </style>
        <div class="table-container"><table class="excel-table">
        <thead>
            <tr><th colspan="13">UPI, Bank & Website Report</th></tr>
            <tr>
                <th rowspan="2">Date</th><th rowspan="2">Total</th>
                <th colspan="5">UPI</th><th colspan="5">Bank</th>
                <th rowspan="2">Unique Website</th>
            </tr>
            <tr>
                <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
                <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
            </tr>
        </thead><tbody>
        """
        for _, row in summary_df.iterrows():
            html_table += f"""<tr>
                <td>{row['Date']}</td><td>{row['Total']}</td>
                <td>{row['UPI_Total']}</td><td>{row['UPI_Unique']}</td><td>{row['UPI_pct']}</td><td>{row['UPI_New']}</td><td>{row['UPI_New_pct']}</td>
                <td>{row['Bank_Total']}</td><td>{row['Bank_Unique']}</td><td>{row['Bank_pct']}</td><td>{row['Bank_New']}</td><td>{row['Bank_New_pct']}</td>
                <td>{row['unique_website']}</td>
            </tr>"""
        html_table += "</tbody></table></div>"
        components.html(html_table, height=450, scrolling=True)

    elif summary_type == "Multiple User's Summary":
        # Show selector for which date to display
        available_dates = sorted(multiple_summary_df["Date"].unique())
        selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        day_df = multiple_summary_df[multiple_summary_df["Date"] == selected_date]

        total_total = day_df["Total"].sum()
        total_unique = day_df["Unique_UPI_Count"].sum()
        total_new = day_df["New_UPI_Count"].sum()
        total_unique_pct = f"{(total_unique / total_total * 100):.0f}%" if total_total else "0%"
        total_new_pct = f"{(total_new / total_unique * 100):.0f}%" if total_unique else "0%"

        table = f"""
        <style>
        .table-user {{ width:100%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
        .table-user th, .table-user td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
        .table-user thead th {{ background:#cfe8b0; font-weight:700; }}
        .table-user tfoot td {{ font-weight:700; background:#cfe8b0; }}
        </style>
        <table class="table-user">
        <thead>
            <tr><th colspan="6">Multiple User's Counts ({selected_date})</th></tr>
            <tr><th rowspan="2">Name</th><th rowspan="2">Total</th><th colspan="2">Unique UPI</th><th colspan="2">New UPI</th></tr>
            <tr><th>Count</th><th>%</th><th>Count</th><th>%</th></tr>
        </thead><tbody>
        """
        for _, row in day_df.iterrows():
            table += f"""<tr>
                <td style="text-align:left">{row['Input_user']}</td>
                <td>{row['Total']}</td><td>{row['Unique_UPI_Count']}</td><td>{row['Unique_UPI_pct']}</td>
                <td>{row['New_UPI_Count']}</td><td>{row['New_UPI_pct']}</td>
            </tr>"""
        table += f"""</tbody>
        <tfoot><tr>
            <td style="text-align:left">Total</td>
            <td>{total_total}</td><td>{total_unique}</td><td>{total_unique_pct}</td>
            <td>{total_new}</td><td>{total_new_pct}</td>
        </tr></tfoot></table>"""
        components.html(table, height=450, scrolling=True)

    elif summary_type == "Employee, Intern & Freelancer Summary":
        available_dates = sorted(freelancer_summary_df["Date"].unique())
        selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        day_df = freelancer_summary_df[freelancer_summary_df["Date"] == selected_date]

        def get_row(user_type):
            r = day_df[day_df["User_Type"] == user_type]
            if r.empty:
                return [0] * 6
            r = r.iloc[0]
            return [r["Total_UPI"], r["Unique_UPI"], r["New_UPI"], r["Total_Bank"], r["Unique_Bank"], r["New_Bank"]]

        e = get_row("Employee")
        i = get_row("INT")
        f = get_row("Freelancer")
        totals = [e[j] + i[j] + f[j] for j in range(6)]

        table = f"""
        <style>
        .table-user {{ width:100%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
        .table-user th, .table-user td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
        .table-user thead th {{ background:#cfe8b0; font-weight:700; }}
        .table-user tfoot td {{ font-weight:700; }}
        </style>
        <table class="table-user">
        <thead>
            <tr><th colspan="8">Employee, Intern & Freelancer Summary ({selected_date})</th></tr>
            <tr><th rowspan="2">User</th><th rowspan="2">Date</th><th colspan="3">UPI</th><th colspan="3">Bank</th></tr>
            <tr><th>Total</th><th>Unique</th><th>New</th><th>Total</th><th>Unique</th><th>New</th></tr>
        </thead><tbody>
            <tr><td>Employee</td><td rowspan="3">{selected_date}</td><td>{e[0]}</td><td>{e[1]}</td><td>{e[2]}</td><td>{e[3]}</td><td>{e[4]}</td><td>{e[5]}</td></tr>
            <tr><td>Intern</td><td>{i[0]}</td><td>{i[1]}</td><td>{i[2]}</td><td>{i[3]}</td><td>{i[4]}</td><td>{i[5]}</td></tr>
            <tr><td>Freelancer</td><td>{f[0]}</td><td>{f[1]}</td><td>{f[2]}</td><td>{f[3]}</td><td>{f[4]}</td><td>{f[5]}</td></tr>
        </tbody>
        <tfoot><tr>
            <td>Total</td><td>NA</td>
            <td>{totals[0]}</td><td>{totals[1]}</td><td>{totals[2]}</td>
            <td>{totals[3]}</td><td>{totals[4]}</td><td>{totals[5]}</td>
        </tr></tfoot></table>"""
        components.html(table, height=450, scrolling=True)

    # FIX 3: Excel export now contains actual data across all 3 sheets
    excel_data = build_excel(summary_df, multiple_summary_df, freelancer_summary_df)

    st.download_button(
        "📥 Download Summary Excel",
        data=excel_data,
        file_name="complete_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("📤 Please upload a file to generate the report.")